import csv
from io import BytesIO, StringIO
from statistics import mean

from django.db.models import Count
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.mentorat.models import (
    MentoreeProgress,
    MentorshipAssignment,
    MentorshipPeriod,
    MentorshipSession,
)


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_CONTENT_TYPE = "text/csv; charset=utf-8"


def display_user(user):
    if not user:
        return "Non renseigne"
    return user.nom_complet or user.email or "Non renseigne"


def display_date(value):
    return value.strftime("%Y-%m-%d") if value else ""


def display_time(value):
    return value.strftime("%H:%M") if value else ""


def display_datetime(value):
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M") if value else ""


def progress_for_assignment(assignment):
    try:
        return assignment.progress
    except MentoreeProgress.DoesNotExist:
        return None


def progress_percentage_for_assignment(assignment):
    progress = progress_for_assignment(assignment)
    if progress and progress.progress_percentage is not None:
        return progress.progress_percentage
    required_sessions = assignment.period.required_sessions or 0
    if not required_sessions:
        return 0
    completed_sessions = assignment.sessions.filter(status=MentorshipSession.Status.COMPLETED).count()
    return round((completed_sessions / required_sessions) * 100)


def get_export_filename(period: MentorshipPeriod, extension: str):
    date_export = timezone.localdate().isoformat()
    slug = slugify(period.title) or f"periode-{period.pk}"
    return f"periode_{slug}_{date_export}.{extension}"


def period_assignments(period: MentorshipPeriod):
    return (
        MentorshipAssignment.objects.filter(period=period)
        .select_related(
            "mentor",
            "mentor__niveau_academique",
            "mentoree",
            "mentoree__niveau_academique",
            "period",
            "progress",
        )
        .prefetch_related("sessions")
        .order_by("mentor__nom", "mentoree__nom", "assigned_at")
    )


def period_sessions(period: MentorshipPeriod):
    return (
        MentorshipSession.objects.filter(assignment__period=period)
        .select_related(
            "assignment",
            "assignment__mentor",
            "assignment__mentor__niveau_academique",
            "assignment__mentoree",
            "assignment__mentoree__niveau_academique",
            "assignment__period",
            "assignment__progress",
        )
        .order_by("assignment__mentor__nom", "assignment__mentoree__nom", "session_number")
    )


def period_statistics(period: MentorshipPeriod):
    assignments = period_assignments(period)
    sessions = period_sessions(period)
    progress_values = [progress_percentage_for_assignment(assignment) for assignment in assignments]
    completed_mentees = sum(
        1
        for assignment in assignments
        if assignment.status == MentorshipAssignment.Status.COMPLETED
        or progress_percentage_for_assignment(assignment) >= 100
    )

    return {
        "Nombre total de mentors": assignments.values("mentor_id").distinct().count(),
        "Nombre total de mentores": assignments.values("mentoree_id").distinct().count(),
        "Nombre total d'affectations": assignments.count(),
        "Nombre total de seances prevues": assignments.count() * period.required_sessions,
        "Nombre total de seances programmees": sessions.filter(status=MentorshipSession.Status.SCHEDULED).count(),
        "Nombre total de seances realisees": sessions.filter(status=MentorshipSession.Status.COMPLETED).count(),
        "Nombre total de seances annulees": sessions.filter(status=MentorshipSession.Status.CANCELLED).count(),
        "Nombre total de seances reportees": sessions.filter(status=MentorshipSession.Status.POSTPONED).count(),
        "Taux moyen d'avancement": round(mean(progress_values), 2) if progress_values else 0,
        "Nombre de mentores termines": completed_mentees,
        "Nombre de mentores encore en cours": max(assignments.count() - completed_mentees, 0),
    }


def write_sheet(sheet, headers, rows):
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F1A17")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append(row)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 45)
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_summary_rows(period: MentorshipPeriod):
    stats = period_statistics(period)
    average_progress = stats["Taux moyen d'avancement"]
    return [
        ["Titre", period.title],
        ["Description", period.description or "Non renseignee"],
        ["Date de debut", display_date(period.start_date)],
        ["Date de fin", display_date(period.end_date)],
        ["Seances obligatoires", period.required_sessions],
        ["Maximum de mentores par mentor", period.max_mentees_per_mentor],
        ["Statut", period.get_status_display()],
        ["Cree le", display_datetime(period.created_at)],
        ["Mis a jour le", display_datetime(period.updated_at)],
        ["Affectations", stats["Nombre total d'affectations"]],
        ["Seances realisees", stats["Nombre total de seances realisees"]],
        ["Taux moyen d'avancement", f"{average_progress}%"],
    ]


def build_mentor_rows(assignments):
    rows = []
    seen = set()
    counts = assignments.values("mentor_id").annotate(total=Count("id"))
    count_by_mentor = {item["mentor_id"]: item["total"] for item in counts}
    for assignment in assignments:
        mentor = assignment.mentor
        if mentor.id in seen:
            continue
        seen.add(mentor.id)
        rows.append(
            [
                display_user(mentor),
                mentor.email,
                mentor.telephone or "",
                mentor.niveau_academique.nom if mentor.niveau_academique else "Non renseigne",
                count_by_mentor.get(mentor.id, 0),
                mentor.statut_compte,
            ]
        )
    return rows


def build_mentoree_rows(assignments):
    rows = []
    seen = set()
    for assignment in assignments:
        mentoree = assignment.mentoree
        if mentoree.id in seen:
            continue
        seen.add(mentoree.id)
        progress = progress_for_assignment(assignment)
        rows.append(
            [
                display_user(mentoree),
                mentoree.email,
                mentoree.telephone or "",
                mentoree.niveau_academique.nom if mentoree.niveau_academique else "Non renseigne",
                assignment.get_status_display(),
                f"{progress_percentage_for_assignment(assignment)}%",
                progress.get_progress_status_display() if progress else "Aucun suivi",
            ]
        )
    return rows


def build_assignment_rows(assignments):
    rows = []
    for assignment in assignments:
        rows.append(
            [
                display_user(assignment.mentor),
                assignment.mentor.email,
                display_user(assignment.mentoree),
                assignment.mentoree.email,
                assignment.get_status_display(),
                display_datetime(assignment.assigned_at),
                assignment.admin_notes or "",
                assignment.sessions.count(),
                assignment.sessions.filter(status=MentorshipSession.Status.COMPLETED).count(),
                f"{progress_percentage_for_assignment(assignment)}%",
            ]
        )
    return rows


def build_session_rows(sessions):
    rows = []
    for session in sessions:
        rows.append(
            [
                display_user(session.assignment.mentor),
                display_user(session.assignment.mentoree),
                session.session_number,
                display_date(session.scheduled_date),
                display_time(session.start_time) or "Non programmee",
                display_time(session.end_time) or "Non programmee",
                session.get_status_display(),
                session.summary or "Non renseigne",
                session.mentor_comment or "Non renseigne",
            ]
        )
    return rows


def build_progress_rows(assignments):
    rows = []
    for assignment in assignments:
        progress = progress_for_assignment(assignment)
        if not progress:
            rows.append(
                [
                    display_user(assignment.mentor),
                    display_user(assignment.mentoree),
                    "Aucun suivi",
                    f"{progress_percentage_for_assignment(assignment)}%",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            continue
        rows.append(
            [
                display_user(assignment.mentor),
                display_user(assignment.mentoree),
                progress.get_progress_status_display(),
                f"{progress.progress_percentage if progress.progress_percentage is not None else progress_percentage_for_assignment(assignment)}%",
                progress.achievements or "Non renseigne",
                progress.difficulties or "Non renseigne",
                progress.recommendations or "Non renseigne",
                progress.mentor_opinion or "Non renseigne",
            ]
        )
    return rows


def build_excel_export(period: MentorshipPeriod):
    assignments = period_assignments(period)
    sessions = period_sessions(period)
    stats = period_statistics(period)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resume periode"
    write_sheet(summary, ["Champ", "Valeur"], build_summary_rows(period))

    mentors = workbook.create_sheet("Mentors")
    write_sheet(
        mentors,
        ["Mentor", "Email", "Telephone", "Niveau academique", "Mentores assignes", "Statut compte"],
        build_mentor_rows(assignments),
    )

    mentorees = workbook.create_sheet("Mentores")
    write_sheet(
        mentorees,
        ["Mentore", "Email", "Telephone", "Niveau academique", "Statut affectation", "Avancement", "Suivi"],
        build_mentoree_rows(assignments),
    )

    assignment_sheet = workbook.create_sheet("Affectations")
    write_sheet(
        assignment_sheet,
        [
            "Mentor",
            "Email mentor",
            "Mentore",
            "Email mentore",
            "Statut",
            "Date affectation",
            "Notes admin",
            "Seances planifiees",
            "Seances realisees",
            "Avancement",
        ],
        build_assignment_rows(assignments),
    )

    sessions_sheet = workbook.create_sheet("Seances")
    write_sheet(
        sessions_sheet,
        [
            "Mentor",
            "Mentore",
            "Numero seance",
            "Date seance",
            "Heure debut",
            "Heure fin",
            "Statut",
            "Resume",
            "Commentaire mentor",
        ],
        build_session_rows(sessions),
    )

    progress_sheet = workbook.create_sheet("Suivis")
    write_sheet(
        progress_sheet,
        [
            "Mentor",
            "Mentore",
            "Appreciation",
            "Avancement",
            "Realisations",
            "Observations",
            "Recommandations",
            "Commentaire mentor",
        ],
        build_progress_rows(assignments),
    )

    stats_sheet = workbook.create_sheet("Statistiques")
    write_sheet(stats_sheet, ["Indicateur", "Valeur"], [[label, value] for label, value in stats.items()])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def csv_rows_for_period(period: MentorshipPeriod):
    assignments = period_assignments(period)
    for assignment in assignments:
        progress = progress_for_assignment(assignment)
        sessions = list(assignment.sessions.all())
        if not sessions:
            yield consolidated_csv_row(period, assignment, None, progress)
            continue
        for session in sessions:
            yield consolidated_csv_row(period, assignment, session, progress)


def consolidated_csv_row(period, assignment, session, progress):
    progress_value = progress.progress_percentage if progress and progress.progress_percentage is not None else progress_percentage_for_assignment(assignment)
    return [
        period.title,
        display_date(period.start_date),
        display_date(period.end_date),
        display_user(assignment.mentor),
        assignment.mentor.email,
        display_user(assignment.mentoree),
        assignment.mentoree.email,
        assignment.mentoree.niveau_academique.nom if assignment.mentoree.niveau_academique else "Non renseigne",
        session.session_number if session else "Non programmee",
        display_date(session.scheduled_date) if session else "Non programmee",
        session.get_status_display() if session else "Non programmee",
        progress.get_progress_status_display() if progress else "Aucun suivi",
        f"{progress_value}%",
        progress.difficulties if progress and progress.difficulties else "Non renseigne",
        progress.recommendations if progress and progress.recommendations else "Non renseigne",
        session.mentor_comment if session and session.mentor_comment else "Non renseigne",
    ]


def build_csv_export(period: MentorshipPeriod):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "periode",
            "date_debut_periode",
            "date_fin_periode",
            "mentor",
            "email_mentor",
            "mentore",
            "email_mentore",
            "niveau_mentore",
            "numero_seance",
            "date_seance",
            "statut_seance",
            "appreciation",
            "avancement",
            "observations",
            "recommandations",
            "commentaire_mentor",
        ]
    )
    writer.writerows(csv_rows_for_period(period))
    return output.getvalue().encode("utf-8-sig")
