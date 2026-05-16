from datetime import date, time
from io import BytesIO

from django.core.exceptions import ValidationError
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from apps.mentorat.models import (
    MentoreeProgress,
    MentorshipAssignment,
    MentorshipPeriod,
    MentorshipPeriodExportLog,
    MentorshipSession,
)
from apps.users.models import NiveauAcademique, Role, Utilisateur


class MentorshipSetupMixin:
    def setUp(self):
        self.role_admin, _ = Role.objects.update_or_create(
            nom=Role.Nom.ADMIN_PRINCIPAL,
            defaults={"description": "Admin"},
        )
        self.role_admin_operationnel, _ = Role.objects.update_or_create(
            nom=Role.Nom.ADMIN_OPERATIONNEL,
            defaults={"description": "Admin operationnel"},
        )
        self.role_mentor, _ = Role.objects.update_or_create(
            nom=Role.Nom.MENTOR,
            defaults={"description": "Mentor"},
        )
        self.role_mentore, _ = Role.objects.update_or_create(
            nom=Role.Nom.MENTORE,
            defaults={"description": "Mentore"},
        )
        self.level_mentoree, _ = NiveauAcademique.objects.update_or_create(
            ordre_niveau=1,
            defaults={
                "nom": "12e annee",
                "est_premier_niveau": True,
                "est_dernier_niveau": False,
            },
        )
        self.level_mentor, _ = NiveauAcademique.objects.update_or_create(
            ordre_niveau=2,
            defaults={
                "nom": "Niveau 1",
                "est_premier_niveau": False,
                "est_dernier_niveau": True,
            },
        )
        self.admin = Utilisateur.objects.create_user(
            email="admin@example.com",
            password="Testpass123!",
            nom="Admin",
            prenom="Awa",
            role=self.role_admin,
            statut_compte=Utilisateur.StatutCompte.ACTIF,
            is_active=True,
            is_staff=True,
            is_superuser=True,
        )
        self.admin_operationnel = Utilisateur.objects.create_user(
            email="admin.operationnel@example.com",
            password="Testpass123!",
            nom="Admin",
            prenom="Oumar",
            role=self.role_admin_operationnel,
            statut_compte=Utilisateur.StatutCompte.ACTIF,
            is_active=True,
            is_staff=True,
        )
        self.mentor = Utilisateur.objects.create_user(
            email="mentor@example.com",
            password="Testpass123!",
            nom="Diop",
            prenom="Aminata",
            role=self.role_mentor,
            profil_mentorat=Utilisateur.ProfilMentorat.MENTOR,
            niveau_academique=self.level_mentor,
            statut_compte=Utilisateur.StatutCompte.ACTIF,
            is_active=True,
            capacite_mentorat=5,
        )
        self.other_mentor = Utilisateur.objects.create_user(
            email="other.mentor@example.com",
            password="Testpass123!",
            nom="Fall",
            prenom="Ibra",
            role=self.role_mentor,
            profil_mentorat=Utilisateur.ProfilMentorat.MENTOR,
            niveau_academique=self.level_mentor,
            statut_compte=Utilisateur.StatutCompte.ACTIF,
            is_active=True,
            capacite_mentorat=5,
        )
        self.mentoree = Utilisateur.objects.create_user(
            email="mentoree@example.com",
            password="Testpass123!",
            nom="Barry",
            prenom="Moussa",
            role=self.role_mentore,
            profil_mentorat=Utilisateur.ProfilMentorat.MENTORE,
            niveau_academique=self.level_mentoree,
            statut_compte=Utilisateur.StatutCompte.ACTIF,
            is_active=True,
        )
        self.period = MentorshipPeriod.objects.create(
            title="Session hiver 2026",
            description="Periode de test",
            start_date=date(2026, 1, 15),
            end_date=date(2026, 4, 30),
            required_sessions=8,
            max_mentees_per_mentor=5,
            status=MentorshipPeriod.Status.ACTIVE,
        )


class MentorshipModelTests(MentorshipSetupMixin, TestCase):
    def test_creation_periode_valide(self):
        period = MentorshipPeriod.objects.create(
            title="Session ete 2026",
            start_date=date(2026, 6, 1),
            end_date=date(2026, 8, 31),
            required_sessions=4,
            status=MentorshipPeriod.Status.DRAFT,
        )

        self.assertEqual(period.required_sessions, 4)
        self.assertEqual(period.max_mentees_per_mentor, 5)
        self.assertEqual(period.status, MentorshipPeriod.Status.DRAFT)

    def test_refus_periode_avec_date_debut_apres_date_fin(self):
        with self.assertRaises(ValidationError):
            MentorshipPeriod.objects.create(
                title="Periode invalide",
                start_date=date(2026, 5, 1),
                end_date=date(2026, 4, 1),
                required_sessions=3,
            )

    def test_refus_mentor_en_12e_annee(self):
        with self.assertRaises(ValidationError):
            Utilisateur.objects.create_user(
                email="mentor.12e@example.com",
                password="Testpass123!",
                nom="Ndiaye",
                prenom="Soda",
                role=self.role_mentor,
                profil_mentorat=Utilisateur.ProfilMentorat.MENTOR,
                niveau_academique=self.level_mentoree,
                statut_compte=Utilisateur.StatutCompte.ACTIF,
                is_active=True,
                capacite_mentorat=1,
            )

    def test_creation_affectation_mentor_mentore(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )

        self.assertEqual(assignment.status, MentorshipAssignment.Status.ACTIVE)
        self.mentor.refresh_from_db()
        self.assertEqual(self.mentor.nombre_mentores_actuels, 1)

    def test_affectation_respecte_limite_de_la_periode(self):
        self.period.max_mentees_per_mentor = 1
        self.period.save(update_fields=["max_mentees_per_mentor", "updated_at"])
        first_mentoree = Utilisateur.objects.create_user(
            email="first.period.limit@example.com",
            password="Testpass123!",
            nom="Limit",
            prenom="First",
            role=self.role_mentore,
            profil_mentorat=Utilisateur.ProfilMentorat.MENTORE,
            niveau_academique=self.level_mentoree,
            statut_compte=Utilisateur.StatutCompte.ACTIF,
            is_active=True,
        )
        second_mentoree = Utilisateur.objects.create_user(
            email="second.period.limit@example.com",
            password="Testpass123!",
            nom="Limit",
            prenom="Second",
            role=self.role_mentore,
            profil_mentorat=Utilisateur.ProfilMentorat.MENTORE,
            niveau_academique=self.level_mentoree,
            statut_compte=Utilisateur.StatutCompte.ACTIF,
            is_active=True,
        )
        MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=first_mentoree,
            period=self.period,
        )

        with self.assertRaises(ValidationError):
            MentorshipAssignment.objects.create(
                mentor=self.mentor,
                mentoree=second_mentoree,
                period=self.period,
            )

    def test_refus_double_affectation_active_meme_periode(self):
        MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )

        with self.assertRaises(ValidationError):
            MentorshipAssignment.objects.create(
                mentor=self.other_mentor,
                mentoree=self.mentoree,
                period=self.period,
            )

    def test_creation_seance_dans_la_periode(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )

        session = MentorshipSession.objects.create(
            assignment=assignment,
            session_number=1,
            scheduled_date=date(2026, 2, 10),
            start_time=time(17, 0),
            end_time=time(18, 0),
        )

        self.assertEqual(session.status, MentorshipSession.Status.SCHEDULED)

    def test_refus_seance_hors_periode(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )

        with self.assertRaises(ValidationError):
            MentorshipSession.objects.create(
                assignment=assignment,
                session_number=1,
                scheduled_date=date(2026, 5, 10),
            )

    def test_refus_numero_seance_superieur_nombre_prevu(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )

        with self.assertRaises(ValidationError):
            MentorshipSession.objects.create(
                assignment=assignment,
                session_number=9,
                scheduled_date=date(2026, 2, 10),
            )


class MentorshipApiTests(MentorshipSetupMixin, APITestCase):
    def test_mise_a_jour_suivi_mentore_par_mentor(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        self.client.force_authenticate(self.mentor)

        response = self.client.patch(
            f"/api/mentor/assignments/{assignment.id}/progress/",
            {
                "progress_status": MentoreeProgress.ProgressStatus.GOOD,
                "progress_percentage": 70,
                "achievements": "Bonne participation.",
                "mentor_opinion": "Progression solide.",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        progress = MentoreeProgress.objects.get(assignment=assignment)
        self.assertEqual(progress.progress_percentage, 70)
        self.assertEqual(progress.progress_status, MentoreeProgress.ProgressStatus.GOOD)

    def test_permissions_admin_et_mentor(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        session = MentorshipSession.objects.create(
            assignment=assignment,
            session_number=1,
            scheduled_date=date(2026, 2, 10),
        )

        response = self.client.get("/api/mentor/dashboard/")
        self.assertIn(response.status_code, [401, 403])

        self.client.force_authenticate(self.admin)
        response = self.client.get("/api/mentorship-periods/")
        self.assertEqual(response.status_code, 200)

        self.client.force_authenticate(self.mentor)
        response = self.client.get("/api/mentorship-periods/")
        self.assertEqual(response.status_code, 403)

        response = self.client.patch(
            f"/api/mentor/sessions/{session.id}/complete/",
            {"summary": "Seance realisee."},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        session.refresh_from_db()
        self.assertEqual(session.status, MentorshipSession.Status.COMPLETED)

        self.client.force_authenticate(self.other_mentor)
        response = self.client.patch(
            f"/api/mentor/sessions/{session.id}/",
            {"summary": "Acces non autorise."},
            format="json",
        )
        self.assertEqual(response.status_code, 404)

    def test_mentor_liste_et_cree_seance_depuis_endpoint_direct(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        self.client.force_authenticate(self.mentor)

        response = self.client.post(
            "/api/mentor/sessions/",
            {
                "assignment": assignment.id,
                "session_number": 1,
                "scheduled_date": "2026-02-10",
                "start_time": "17:00",
                "end_time": "18:00",
                "summary": "Objectifs de debut de periode.",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        response = self.client.get("/api/mentor/sessions/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)

    def test_mentor_follow_up_complete_seance_et_genere_progression(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        MentorshipSession.objects.create(
            assignment=assignment,
            session_number=1,
            scheduled_date=date(2026, 2, 3),
            status=MentorshipSession.Status.COMPLETED,
        )
        session = MentorshipSession.objects.create(
            assignment=assignment,
            session_number=2,
            scheduled_date=date(2026, 2, 10),
        )
        self.client.force_authenticate(self.mentor)

        response = self.client.patch(
            f"/api/mentor/follow-ups/{session.id}/",
            {
                "progress_status": MentoreeProgress.ProgressStatus.GOOD,
                "appreciation": "Tres bon",
                "observation": "Bonne participation.",
                "recommendations": "Continuer les exercices.",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        session.refresh_from_db()
        self.assertEqual(session.status, MentorshipSession.Status.COMPLETED)
        progress = MentoreeProgress.objects.get(assignment=assignment)
        self.assertEqual(progress.progress_percentage, 25)
        self.assertEqual(progress.progress_status, MentoreeProgress.ProgressStatus.GOOD)

    def test_ancien_endpoint_disponibilites_est_desactive(self):
        response = self.client.get(f"/api/mentors/{self.mentor.id}/available-slots/")

        self.assertEqual(response.status_code, 410)

    def test_mentor_peut_reconduire_affectation_sur_nouvelle_periode(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        next_period = MentorshipPeriod.objects.create(
            title="Session 2027",
            start_date=date(2027, 1, 15),
            end_date=date(2027, 4, 30),
            required_sessions=6,
            status=MentorshipPeriod.Status.DRAFT,
        )
        self.client.force_authenticate(self.mentor)

        response = self.client.post(
            f"/api/mentor/assignments/{assignment.id}/continue/",
            {"period": next_period.id},
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertTrue(
            MentorshipAssignment.objects.filter(
                mentor=self.mentor,
                mentoree=self.mentoree,
                period=next_period,
                status=MentorshipAssignment.Status.ACTIVE,
            ).exists()
        )

    def test_inscription_mentor_refuse_12e_annee(self):
        response = self.client.post(
            "/api/inscriptions/mentor/",
            {
                "nom": "Diallo",
                "prenom": "Mariama",
                "email": "mentor.12e.inscription@example.com",
                "telephone": "",
                "langue_preferee": "FR",
                "region": "",
                "niveau_academique": self.level_mentoree.id,
                "mentorship_period": self.period.id,
                "mini_bio": "Je souhaite accompagner la releve.",
                "capacite_mentorat": 1,
                "consentement": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("niveau_academique", response.data)

    def test_profil_mentor_public_requiert_mini_bio(self):
        self.client.force_authenticate(self.mentor)

        response = self.client.patch(
            "/api/mentor/profile/",
            {"wants_to_appear_on_team_page": True, "mini_bio": ""},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("mini_bio", response.data)
        self.assertIn("domaine_specialite", response.data)
        self.assertIn("profile_photo", response.data)

    def test_profil_mentor_ne_modifie_pas_le_niveau_academique(self):
        self.mentor.profile_photo = "mentor_profiles/mentor.jpg"
        self.mentor.save(update_fields=["profile_photo"])
        self.client.force_authenticate(self.mentor)

        response = self.client.patch(
            "/api/mentor/profile/",
            {
                "niveau_academique": self.level_mentoree.id,
                "mini_bio": "Mentore engagee.",
                "domaine_specialite": "Sciences de la sante",
                "wants_to_appear_on_team_page": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.mentor.refresh_from_db()
        self.assertEqual(self.mentor.niveau_academique, self.level_mentor)

    def test_admin_export_excel_periode_reussi(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        MentorshipSession.objects.create(
            assignment=assignment,
            session_number=1,
            scheduled_date=date(2026, 2, 10),
            status=MentorshipSession.Status.COMPLETED,
            summary="Seance realisee.",
            mentor_comment="Bonne progression.",
        )
        MentoreeProgress.objects.create(
            assignment=assignment,
            progress_status=MentoreeProgress.ProgressStatus.GOOD,
            progress_percentage=75,
            recommendations="Continuer les exercices.",
        )
        self.client.force_authenticate(self.admin)

        response = self.client.get(f"/api/admin/mentorship-periods/{self.period.id}/export/excel/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("periode_session-hiver-2026_", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Resume periode", "Mentors", "Mentores", "Affectations", "Seances", "Suivis", "Statistiques"],
        )
        self.assertEqual(workbook["Resume periode"]["B2"].value, "Session hiver 2026")
        self.assertEqual(workbook["Mentors"]["A2"].value, self.mentor.nom_complet)
        self.assertTrue(
            MentorshipPeriodExportLog.objects.filter(
                period=self.period,
                format=MentorshipPeriodExportLog.Format.EXCEL,
                exported_by=self.admin,
            ).exists()
        )

    def test_admin_export_csv_periode_reussi(self):
        assignment = MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        MentorshipSession.objects.create(
            assignment=assignment,
            session_number=1,
            scheduled_date=date(2026, 2, 10),
            status=MentorshipSession.Status.COMPLETED,
        )
        self.client.force_authenticate(self.admin)

        response = self.client.get(f"/api/admin/mentorship-periods/{self.period.id}/export/csv/")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("text/csv"))
        csv_content = response.content.decode("utf-8-sig")
        self.assertIn("Session hiver 2026", csv_content)
        self.assertIn("mentor@example.com", csv_content)
        self.assertIn("mentoree@example.com", csv_content)
        self.assertTrue(
            MentorshipPeriodExportLog.objects.filter(
                period=self.period,
                format=MentorshipPeriodExportLog.Format.CSV,
                exported_by=self.admin,
            ).exists()
        )

    def test_export_refuse_aux_non_admins(self):
        self.client.force_authenticate(self.mentor)

        response = self.client.get(f"/api/admin/mentorship-periods/{self.period.id}/export/csv/")

        self.assertEqual(response.status_code, 403)

    def test_export_refuse_admin_operationnel(self):
        self.client.force_authenticate(self.admin_operationnel)

        response = self.client.get(f"/api/admin/mentorship-periods/{self.period.id}/export/csv/")

        self.assertEqual(response.status_code, 403)
        self.assertFalse(
            MentorshipPeriodExportLog.objects.filter(
                period=self.period,
                exported_by=self.admin_operationnel,
            ).exists()
        )

    def test_export_periode_inexistante(self):
        self.client.force_authenticate(self.admin)

        response = self.client.get("/api/admin/mentorship-periods/99999/export/excel/")

        self.assertEqual(response.status_code, 404)

    def test_export_csv_fonctionne_avec_donnees_incompletes(self):
        MentorshipAssignment.objects.create(
            mentor=self.mentor,
            mentoree=self.mentoree,
            period=self.period,
        )
        self.client.force_authenticate(self.admin)

        response = self.client.get(f"/api/admin/mentorship-periods/{self.period.id}/export/csv/")

        self.assertEqual(response.status_code, 200)
        csv_content = response.content.decode("utf-8-sig")
        self.assertIn("Non programmee", csv_content)
        self.assertIn("Aucun suivi", csv_content)

    def test_admin_valide_affichage_equipe_et_api_publique_filtre(self):
        self.mentor.mini_bio = "Mentore engagee en sciences de la sante."
        self.mentor.domaine_specialite = "Sciences de la sante"
        self.mentor.profile_photo = "mentor_profiles/mentor.jpg"
        self.mentor.wants_to_appear_on_team_page = True
        self.mentor.save(
            update_fields=[
                "mini_bio",
                "domaine_specialite",
                "profile_photo",
                "wants_to_appear_on_team_page",
            ]
        )

        response = self.client.get("/api/public/team/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, [])

        self.client.force_authenticate(self.admin)
        response = self.client.patch(
            f"/api/admin/team-members/{self.mentor.id}/",
            {"is_team_approved": True, "team_display_order": 1},
            format="json",
        )
        self.assertEqual(response.status_code, 200)

        self.client.force_authenticate(user=None)
        response = self.client.get("/api/public/team/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["nom_complet"], self.mentor.nom_complet)

    def test_admin_refuse_approbation_equipe_sans_ordre_positif(self):
        self.mentor.mini_bio = "Mentore engagee."
        self.mentor.wants_to_appear_on_team_page = True
        self.mentor.save(update_fields=["mini_bio", "wants_to_appear_on_team_page"])
        self.client.force_authenticate(self.admin)

        response = self.client.patch(
            f"/api/admin/team-members/{self.mentor.id}/",
            {"is_team_approved": True, "team_display_order": 0},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("team_display_order", response.data)

    def test_inscription_mentore_refuse_niveau_medecine(self):
        medecine, _ = NiveauAcademique.objects.update_or_create(
            ordre_niveau=4,
            defaults={
                "nom": "Je suis etudiant(e) en medecine",
                "est_premier_niveau": False,
                "est_dernier_niveau": True,
            },
        )

        response = self.client.post(
            "/api/inscriptions/mentore/",
            {
                "nom": "Kane",
                "prenom": "Fatou",
                "email": "mentore.medecine@example.com",
                "telephone": "",
                "langue_preferee": "FR",
                "region": "",
                "niveau_academique": medecine.id,
                "mentorship_period": self.period.id,
                "objectifs": "Trouver un mentor.",
                "mentor_choisi": self.mentor.id,
                "consentement": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("niveau_academique", response.data)
